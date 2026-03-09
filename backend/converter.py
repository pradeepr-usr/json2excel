# # # converter.py

from openpyxl import Workbook
from io import BytesIO
import time
import json
from openpyxl.styles import Font


class SheetManager:

    def __init__(self, workbook):
        self.workbook = workbook
        self.sheets = {}
        self.headers = {}
        self.sheet_name_map = {}
        self.sheet_name_counter = {}
        self.column_widths = {}
        self.row_counts = {}

    def clean_sheet_name(self, name):
        base = name.split("_")[-1]
        if base not in self.sheet_name_counter:
            self.sheet_name_counter[base] = 1
            return base[:31]
        self.sheet_name_counter[base] += 1
        new_name = f"{base}_{self.sheet_name_counter[base]}"
        return new_name[:31]

    def get_sheet(self, raw_name):
        if raw_name not in self.sheet_name_map:
            clean = self.clean_sheet_name(raw_name)
            self.sheet_name_map[raw_name] = clean

        sheet_name = self.sheet_name_map[raw_name]

        if sheet_name not in self.sheets:
            if sheet_name == "root":
                sheet = self.workbook.active
                sheet.title = "root"
            else:
                sheet = self.workbook.create_sheet(sheet_name)

            self.sheets[sheet_name] = sheet
            self.headers[sheet_name] = None
            self.column_widths[sheet_name] = {}
            self.row_counts[sheet_name] = 0

        return sheet_name, self.sheets[sheet_name]

    def update_column_width(self, sheet_name, column_index, value):
        value_length = len(str(value)) if value is not None else 0
        current = self.column_widths[sheet_name].get(column_index, 0)
        if value_length > current:
            self.column_widths[sheet_name][column_index] = value_length

    def write_row(self, raw_name, row):
        sheet_name, sheet = self.get_sheet(raw_name)

        if self.headers[sheet_name] is None:
            headers = list(row.keys())
            self.headers[sheet_name] = headers
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            for idx, header in enumerate(headers, start=1):
                self.update_column_width(sheet_name, idx, header)

        headers = self.headers[sheet_name]
        values = [row.get(h) for h in headers]
        sheet.append(values)
        self.row_counts[sheet_name] += 1

        for idx, value in enumerate(values, start=1):
            self.update_column_width(sheet_name, idx, value)

    def apply_column_widths(self):
        for sheet_name, sheet in self.sheets.items():
            widths = self.column_widths[sheet_name]
            for col_idx, width in widths.items():
                col_letter = sheet.cell(row=1, column=col_idx).column_letter
                sheet.column_dimensions[col_letter].width = min(width + 2, 50)

    def get_total_rows(self):
        return sum(self.row_counts.values())


def extract_fields(data) -> list[str]:
    fields = []
    seen = set()

    def walk(node, prefix=""):
        if isinstance(node, dict):
            for k, v in node.items():
                path = f"{prefix}.{k}" if prefix else k
                if isinstance(v, dict):
                    walk(v, path)
                elif isinstance(v, list) and v and isinstance(v[0], dict):
                    walk(v[0], f"{path}[]")
                else:
                    if path not in seen:
                        seen.add(path)
                        fields.append(path)
        elif isinstance(node, list) and node:
            walk(node[0], prefix)

    walk(data)
    return fields


def json_to_excel_bytes(data, selected_fields: list[str] | None = None):
    """Single file conversion — returns (buffer, summary)."""
    start_time = time.time()

    filter_active = selected_fields is not None and len(selected_fields) > 0

    workbook = Workbook()
    sheet_manager = SheetManager(workbook)

    global_id = 1

    def generate_id():
        nonlocal global_id
        current = global_id
        global_id += 1
        return current

    def flatten_dict(d, parent_key=""):
        items = {}
        stack = [(d, parent_key)]
        while stack:
            current_dict, prefix = stack.pop()
            for k, v in current_dict.items():
                new_key = f"{prefix}.{k}" if prefix else k
                if isinstance(v, dict):
                    stack.append((v, new_key))
                else:
                    items[new_key] = v
        return items

    def should_include_field(field_name: str, sheet_name: str) -> bool:
        if not filter_active:
            return True
        if field_name in ("_id", "_parent_id"):
            return True
        if sheet_name == "root":
            return field_name in selected_fields
        sheet_key = sheet_name.split("_")[-1]
        qualified = f"{sheet_key}[].{field_name}"
        return qualified in selected_fields

    def process_node(node, sheet_name="root", parent_id=None):
        if isinstance(node, list):
            for item in node:
                process_node(item, sheet_name, parent_id)
            return

        if isinstance(node, dict):
            row = {}
            current_id = generate_id()
            row["_id"] = current_id
            row["_parent_id"] = parent_id

            for key, value in node.items():
                if isinstance(value, dict):
                    flat = flatten_dict(value, key)
                    for flat_key, flat_val in flat.items():
                        if should_include_field(flat_key, sheet_name):
                            row[flat_key] = flat_val
                elif isinstance(value, list):
                    if not value:
                        continue
                    child_sheet = f"{sheet_name}_{key}"
                    if all(isinstance(i, dict) for i in value):
                        for item in value:
                            process_node(item, child_sheet, current_id)
                    else:
                        for item in value:
                            primitive_row = {
                                "_id": generate_id(),
                                "_parent_id": current_id,
                                "value": item
                            }
                            sheet_manager.write_row(child_sheet, primitive_row)
                else:
                    if should_include_field(key, sheet_name):
                        row[key] = value

            sheet_manager.write_row(sheet_name, row)

    if isinstance(data, list):
        process_node(data, "root", None)
    elif isinstance(data, dict):
        process_node(data, "root", None)
    else:
        raise ValueError("Unsupported JSON structure")

    sheet_manager.apply_column_widths()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    end_time = time.time()
    file_size_bytes = output.getbuffer().nbytes

    summary = {
        "sheets": len(sheet_manager.sheets),
        "rows": sheet_manager.get_total_rows(),
        "file_size_kb": round(file_size_bytes / 1024, 1),
        "time_sec": round(end_time - start_time, 3),
    }

    print("------ JSON → Excel Report ------")
    print(f"Sheets   : {summary['sheets']}")
    print(f"Rows     : {summary['rows']}")
    print(f"Size     : {summary['file_size_kb']} KB")
    print(f"Time     : {summary['time_sec']}s")
    print(f"Filter   : {'active' if filter_active else 'off'}")
    print("---------------------------------")

    return output, summary


def batch_to_excel_bytes(files: list[tuple[str, bytes]]):
    """
    Batch conversion — multiple JSON files into one Excel workbook.
    Each file becomes one or more sheets named after the filename.
    files: list of (filename_without_ext, raw_bytes)
    Returns (buffer, summary)
    """
    start_time = time.time()

    workbook = Workbook()
    # Remove default empty sheet — we'll create named ones
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    total_rows = 0
    total_sheets = 0
    file_results = []

    global_id = 1

    def generate_id():
        nonlocal global_id
        current = global_id
        global_id += 1
        return current

    def flatten_dict(d, parent_key=""):
        items = {}
        stack = [(d, parent_key)]
        while stack:
            current_dict, prefix = stack.pop()
            for k, v in current_dict.items():
                new_key = f"{prefix}.{k}" if prefix else k
                if isinstance(v, dict):
                    stack.append((v, new_key))
                else:
                    items[new_key] = v
        return items

    # Track used sheet names across all files to avoid collisions
    used_sheet_names: set[str] = set()

    def safe_sheet_name(name: str) -> str:
        base = name[:31]
        if base not in used_sheet_names:
            used_sheet_names.add(base)
            return base
        i = 2
        while True:
            candidate = f"{name[:28]}_{i}"[:31]
            if candidate not in used_sheet_names:
                used_sheet_names.add(candidate)
                return candidate
            i += 1

    def process_file(file_name: str, data):
        """Process one JSON file into the shared workbook."""
        nonlocal total_rows, total_sheets, global_id

        # Per-file sheet manager but using shared workbook
        sheets: dict[str, object] = {}
        headers: dict[str, list] = {}
        column_widths: dict[str, dict] = {}
        row_counts: dict[str, int] = {}

        # Map internal sheet key → actual Excel sheet name
        # Root sheet for this file is named after the file
        sheet_name_map: dict[str, str] = {}

        def get_or_create_sheet(raw_key: str) -> tuple[str, object]:
            if raw_key in sheet_name_map:
                sheet_name = sheet_name_map[raw_key]
                return sheet_name, sheets[sheet_name]

            # Derive display name
            if raw_key == "root":
                display = safe_sheet_name(file_name)
            else:
                # e.g. "root_items" → "orders_items"
                suffix = raw_key[len("root"):]  # "_items"
                display = safe_sheet_name(f"{file_name}{suffix}")

            sheet_name_map[raw_key] = display
            sheet = workbook.create_sheet(display)
            sheets[display] = sheet
            headers[display] = None
            column_widths[display] = {}
            row_counts[display] = 0
            return display, sheet

        def update_col_width(sname, col_idx, value):
            vl = len(str(value)) if value is not None else 0
            current = column_widths[sname].get(col_idx, 0)
            if vl > current:
                column_widths[sname][col_idx] = vl

        def write_row(raw_key: str, row: dict):
            sname, sheet = get_or_create_sheet(raw_key)
            if headers[sname] is None:
                hdrs = list(row.keys())
                headers[sname] = hdrs
                sheet.append(hdrs)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                for idx, h in enumerate(hdrs, start=1):
                    update_col_width(sname, idx, h)
            hdrs = headers[sname]
            values = [row.get(h) for h in hdrs]
            sheet.append(values)
            row_counts[sname] += 1
            for idx, v in enumerate(values, start=1):
                update_col_width(sname, idx, v)

        def process_node(node, sheet_key="root", parent_id=None):
            if isinstance(node, list):
                for item in node:
                    process_node(item, sheet_key, parent_id)
                return
            if isinstance(node, dict):
                row = {}
                current_id = generate_id()
                row["_id"] = current_id
                row["_parent_id"] = parent_id
                for key, value in node.items():
                    if isinstance(value, dict):
                        flat = flatten_dict(value, key)
                        row.update(flat)
                    elif isinstance(value, list):
                        if not value:
                            continue
                        child_key = f"{sheet_key}_{key}"
                        if all(isinstance(i, dict) for i in value):
                            for item in value:
                                process_node(item, child_key, current_id)
                        else:
                            for item in value:
                                write_row(child_key, {
                                    "_id": generate_id(),
                                    "_parent_id": current_id,
                                    "value": item
                                })
                    else:
                        row[key] = value
                write_row(sheet_key, row)

        if isinstance(data, list):
            process_node(data, "root", None)
        elif isinstance(data, dict):
            process_node(data, "root", None)

        # Apply column widths for this file's sheets
        for sname, sheet in sheets.items():
            for col_idx, width in column_widths[sname].items():
                col_letter = sheet.cell(row=1, column=col_idx).column_letter
                sheet.column_dimensions[col_letter].width = min(width + 2, 50)

        file_rows = sum(row_counts.values())
        file_sheets = len(sheets)
        total_rows += file_rows
        total_sheets += file_sheets

        file_results.append({
            "name": file_name,
            "sheets": file_sheets,
            "rows": file_rows,
        })

    # Process each file
    errors = []
    for file_name, raw_bytes in files:
        try:
            data = json.loads(raw_bytes)
            process_file(file_name, data)
        except Exception as e:
            print(f"ERROR processing {file_name}: {e}")  # ADD THIS
            import traceback; traceback.print_exc()       # ADD THIS
            errors.append({"name": file_name, "error": str(e)})

    if total_sheets == 0:
        raise ValueError("No valid JSON files could be processed")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    end_time = time.time()
    file_size_bytes = output.getbuffer().nbytes

    summary = {
        "files": len(file_results),
        "sheets": total_sheets,
        "rows": total_rows,
        "file_size_kb": round(file_size_bytes / 1024, 1),
        "time_sec": round(end_time - start_time, 3),
        "errors": errors,
    }

    print("------ Batch JSON → Excel Report ------")
    print(f"Files    : {summary['files']}")
    print(f"Sheets   : {summary['sheets']}")
    print(f"Rows     : {summary['rows']}")
    print(f"Size     : {summary['file_size_kb']} KB")
    print(f"Time     : {summary['time_sec']}s")
    if errors:
        print(f"Errors   : {len(errors)}")
    print("---------------------------------------")

    return output, summary